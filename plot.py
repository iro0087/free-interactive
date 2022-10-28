#!usr/bin python3 

import matplotlib

matplotlib.use("TkAgg")

import matplotlib.pyplot as plt

from openpyxl import load_workbook

import os

from tkinter import Tk, Label, Button, Entry

import sys

lf = os.listdir(".")

nbrstr = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]

def quit():

    sys.exit()

def fun():

    w = load_workbook("data.xlsx")

    sheet = w.active

    fichier = varfich.get()

    if lf.count(fichier + ".png") > 0:

        fichier = fichier + ".png"

    if lf.count(fichier + ".jpg") > 0:

        fichier = fichier + ".jpg"

    if lf.count(fichier) == 0:

        erf = Label(fen, text="file does not exist", fg="red")

        erf.pack()

    titre = vartitre.get()

    titrey = vary.get()

    titrex = varx.get()

    gg = vargrid.get()

    ygv = varm.get()

    if len(xmax.get()) > 0:

        xmax2 = float(xmax.get())

    if len(ymax.get()):

        ymax2 = float(ymax.get())

    if ygv != "w" and ygv != "m":

        Label(fen, text="Just put 'm' or 'w'", fg="red").pack()

    t = 2

    t2 = 1

    lx = []

    passe = 0

    ly = []

    ly2 = []

    lydemo = []

    colorl = []

    while sheet.cell(row=t, column=1).value != None and passe == 0:

        if sheet.cell(row=t, column=1).value == "#y":

            passe = 1

        else:

            t += 1

    numx = t

    t = 1

    passe = 0

    while sheet.cell(row=1, column=t2).value != None or t2 == 1:

        if t2 == 1 and t > 1:

            if sheet.cell(row=t, column=t2).value != None and t < numx:

                lx.append(sheet.cell(row=t, column=t2).value)

            if sheet.cell(row=t, column=t2).value != None and t > numx:

                lydemo.append(sheet.cell(row=t, column=t2).value)

        if t2 > 1:

            if t > 1 and t < numx and sheet.cell(row=t, column=t2).value != None and passe == 0:

                ly.append(sheet.cell(row=t, column=t2).value)

                passe = 1

            else:

                if sheet.cell(row=t, column=t2).value != None and t > 1:

                    colorl.append(sheet.cell(row=t, column=t2).value)
                    
                    t2 += 1

                    t = 1

        if sheet.cell(row=t, column=t2).value == None:

            if sheet.cell(row=t, column=1).value != None and t2 > 1 and passe == 0 and t > 1 and t < numx: 

                ly.append("#")

                passe = 1

            else:

                if t2 == 1:

                    t2 += 1

                    t = 1

        passe = 0

        t += 1

    really = len(ly)

    t2 = 0

    lx2 = []

    lx3 = []

    t = 0

    while t2 < len(ly):

        if ly[t2] != "#":

            lx2.append(lx[t])

        if t + 1 == len(lx):

            lx3.append(lx2)

            lx2 = []

            t = -1

        t += 1

        t2 += 1

    t2 = 0

    t = 0

    while t2 < len(ly):

        if ly[t2] == "#":

            ly.pop(t2)

            t2 = -1

        t2 += 1

    maxy = 0
    
    t = 0

    if len(lydemo) == 0:

        r_max2 = ymax2

        if len(xmax.get()) > 0:

            r_max1 = xmax2

        else:

            r_max1 = len(lx) - 1

    else:

        r_max2 = len(lydemo) - 1

        if len(xmax.get()) > 0:

            r_max1 = xmax2

        else:

            r_max1 = len(lx) - 1

    img = plt.imread(fichier)

    fig, ax = plt.subplots()

    ax.imshow(img, extent=[0, r_max1, 0, r_max2])

    t2 = 0

    lyref = []

    if nbrstr.count(lx) == 0:

        while t2 < really / len(colorl):   

            lyref.append(0)

            t2 += 1

        plt.scatter(lx, lyref, marker="", color=colorl[t])

    else:

        while t2 < len(lx):

            lx[t2] = float(lx[t2])

            t2 += 1

    t2 = 0

    lxref = []

    if len(lydemo) > 0:

        while t2 < len(lydemo):   

            lxref.append(0)

            t2 += 1

        plt.scatter(lxref, lydemo, marker="", color=colorl[t])

    t2 = 0

    if gg == "y":

        plt.grid()

    pdp = 0

    t2 = 0

    while t < really / len(lx):

        ly2 = []

        while t2 < len(lx3[t]):

            ly2.append(ly[pdp + t2])

            t2 += 1

        t2 = 0

        phry = sheet.cell(row=1, column=t + 2).value

        while t2 < len(ly2):

            if ygv == "w":

                plt.text(lx3[t][t2], ly2[t2], phry, rotation=90, color=colorl[t], clip_on=True)

            else:

                if ygv == "m":

                    plt.text(lx3[t][t2], ly2[t2], "*", color=colorl[t], clip_on=True)

            t2 += 1

        t2 = 0

        pdp = pdp + len(lx3[t])

        t += 1

    t = 0

    t2 = 0

    while t < len(lx):

        yval = sheet.cell(row=1, column=t + 1).value

        t += 1

    plt.title(titre)

    plt.ylabel(titrey)

    plt.xlabel(titrex)

    ax.set_facecolor("grey")

    plt.show()

fen = Tk()

Label(fen, text="").pack()

label = Label(fen, text="Please enter the name of the image file")

label1 = Label(fen, text="What is the name of the graph? (skippable)")

label2 = Label(fen, text="Name of y ?(skippable)")

label3 = Label(fen, text="Name of x ?(skippable)")

label4 = Label(fen, text="Put the grid? May increase accuracy. (y/n)")

label5 = Label(fen, text="Do you want to display markers or words? (m/w)")

label.pack()

varfich = Entry(fen, width=30)

vartitre = Entry(fen, width=30)

vary = Entry(fen, width=30)

varx = Entry(fen, width=30)

vargrid = Entry(fen, width=30)

varm = Entry(fen, width=30)

varfich.pack()

label1.pack()

vartitre.pack()

label2.pack()

vary.pack()

label3.pack()

varx.pack()

label4.pack()

vargrid.pack()

label5.pack()

varm.pack()

Label(fen, text="What is xmax? (skippable if not number)").pack()

xmax = Entry(fen, width=30)

xmax.pack()

Label(fen, text="What is ymax? (skippable if not number)").pack()

ymax = Entry(fen, width=30)

ymax.pack()

Button(fen, text="PROCEED", command=fun, bg="yellow").pack()

Label(fen, text="").pack()

Button(fen, text="TERMINATE", command=quit, bg="red").pack()

fen.mainloop()
